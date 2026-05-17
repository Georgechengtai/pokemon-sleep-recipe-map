#!/usr/bin/env python3
"""
prep_data.py — Excel → data.json

Reads Pokemon Sleep Recipe.xlsx and outputs data.json with:
  - ingredients: list of 19 ingredients
  - recipes: list of 79 recipes with ingredient breakdown
  - incidence_matrix: 19×N binary matrix (ingredient i used by recipe j)
  - cooccurrence: 19×19 symmetric count matrix

Usage:
    python3 scripts/prep_data.py [--out data.json]
"""

import json
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: pip3 install openpyxl")
    sys.exit(1)

try:
    import numpy as np
    HAS_NUMPY = True
except ImportError:
    HAS_NUMPY = False

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "Pokemon Sleep Recipe.xlsx"
OUT  = ROOT / "data.json"


# Local asset filename → ingredient English name (lowercase, no space)
LOCAL_ASSETS = {
    "bean_sausage":        "beansausage",
    "fancy_apple":         "fancyapple",
    "fancy_egg":           "fancyegg",
    "fiery_herb":          "fieryherb",
    "greengrass_soybeans": "greengrasssoybeans",
    "honey":               "honey",
    "large_leek":          "largeleek",
    "moomoo_milk":         "moomoomilk",
    "pure_oil":            "pureoil",
    "slowpoke_tail":       "slowpoketail",
    "snoozy_tomato":       "snoozytomato",
    "soft_potato":         "softpotato",
    "soothing_cacao":      "soothingcacao",
    "tasty_mushroom":      "tastymushroom",
    "warming_ginger":      "warmingginger",
}
# Build reverse map: key_no_space → local filename stem
ASSET_MAP = {v: k for k, v in LOCAL_ASSETS.items()}


def serebii_ingredient_url(name_en: str) -> str:
    key = name_en.lower().replace(" ", "")
    return f"https://www.serebii.net/pokemonsleep/ingredients/{key}.png"


def serebii_recipe_url(name_en_code: str) -> str:
    return f"https://www.serebii.net/pokemonsleep/meals/{name_en_code}.png"


def local_ingredient_path(name_en: str) -> str | None:
    key = name_en.lower().replace(" ", "")
    stem = ASSET_MAP.get(key)
    return f"assets/{stem}.jpg" if stem else None


def parse_ingredients(wb) -> list[dict]:
    ws = wb["食材"]
    ingredients = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        id_, name, name_en, _img, init_level, base_score, release_date = row[:7]
        if not id_:
            continue
        name_en = name_en or ""
        key = name_en.lower().replace(" ", "")
        entry = {
            "id":           int(id_),
            "name":         name,
            "name_en":      name_en,
            "base_score":   int(base_score) if base_score else 0,
            "init_level":   int(init_level) if init_level else 1,
            "release_date": release_date.strftime("%Y-%m-%d") if isinstance(release_date, datetime) else str(release_date),
            "img_serebii":  serebii_ingredient_url(name_en),
            "img_local":    local_ingredient_path(name_en),
            "key":          key,
        }
        ingredients.append(entry)
    return ingredients


def parse_recipes(wb) -> list[dict]:
    ws = wb["食譜"]
    # Row 2 is the header: 圖,食譜名,名,食材數量,類型,階級獎勵,...
    # Data starts at row 3
    recipes = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[1]
        if not name:
            continue
        name_en_code = row[2] or ""
        ing_count    = int(row[3]) if row[3] else 0
        type_        = row[4] or ""
        tier_bonus   = float(row[5]) if row[5] else 0.0
        score_base   = int(row[6]) if row[6] else 0
        score_final  = int(row[7]) if row[7] else 0

        # Ingredient pairs: cols J/K, L/M, N/O, P/Q (indices 9-16)
        ingredients = []
        for i in range(4):
            ing_name  = row[9 + i * 2]
            ing_count_ = row[10 + i * 2]
            if ing_name and ing_count_:
                ingredients.append({
                    "name":  ing_name,
                    "count": int(ing_count_),
                })

        # Normalise type label
        type_map = {
            "咖哩、濃湯": "curry",
            "沙拉":        "salad",
            "點心、飲料":  "dessert",
        }
        type_en = type_map.get(type_, type_)

        recipes.append({
            "name":        name,
            "name_en":     name_en_code,
            "type":        type_,
            "type_en":     type_en,
            "tier_bonus":  tier_bonus,
            "score_base":  score_base,
            "score_final": score_final,
            "ingredients": ingredients,
            "img_serebii": serebii_recipe_url(name_en_code),
        })

    return recipes


def build_incidence(ingredients: list[dict], recipes: list[dict]) -> list[list[int]]:
    """Returns 19×N binary matrix: M[i][j] = 1 if ingredient i used by recipe j."""
    name_to_idx = {ing["name"]: i for i, ing in enumerate(ingredients)}
    n_ing = len(ingredients)
    n_rec = len(recipes)
    M = [[0] * n_rec for _ in range(n_ing)]
    for j, recipe in enumerate(recipes):
        for ing in recipe["ingredients"]:
            i = name_to_idx.get(ing["name"])
            if i is not None:
                M[i][j] = 1
    return M


def build_cooccurrence(ingredients: list[dict], M: list[list[int]]) -> list[list[int]]:
    """Returns 19×19 co-occurrence count matrix C = M·Mᵀ."""
    n = len(ingredients)
    n_rec = len(M[0])
    C = [[0] * n for _ in range(n)]
    for i in range(n):
        for j in range(n):
            C[i][j] = sum(M[i][k] * M[j][k] for k in range(n_rec))
    return C


def spectral_bicluster_order(M: list[list[int]]) -> tuple[list[int], list[int]]:
    """
    Returns (row_order, col_order) using SVD of normalized incidence matrix.
    Sorts by 2nd singular vector (Fiedler-like ordering).
    Falls back to frequency-based sort if numpy unavailable.
    """
    n_row = len(M)
    n_col = len(M[0])

    if HAS_NUMPY:
        A = np.array(M, dtype=float)
        row_sums = A.sum(axis=1)
        col_sums = A.sum(axis=0)
        # Avoid division by zero
        row_sums = np.where(row_sums == 0, 1, row_sums)
        col_sums = np.where(col_sums == 0, 1, col_sums)
        Dr_inv_sqrt = np.diag(1.0 / np.sqrt(row_sums))
        Dc_inv_sqrt = np.diag(1.0 / np.sqrt(col_sums))
        A_norm = Dr_inv_sqrt @ A @ Dc_inv_sqrt
        U, S, Vt = np.linalg.svd(A_norm, full_matrices=False)
        # Use 2nd singular vector (index 1) for ordering
        row_order = np.argsort(U[:, 1]).tolist()
        col_order = np.argsort(Vt[1, :]).tolist()
    else:
        # Fallback: sort rows by total frequency desc, cols by type then tier
        row_sums = [sum(M[i]) for i in range(n_row)]
        row_order = sorted(range(n_row), key=lambda i: -row_sums[i])
        col_sums = [sum(M[i][j] for i in range(n_row)) for j in range(n_col)]
        col_order = sorted(range(n_col), key=lambda j: -col_sums[j])

    return row_order, col_order


def main(out_path: Path = OUT):
    print(f"Loading {XLSX}...")
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    ingredients = parse_ingredients(wb)
    recipes     = parse_recipes(wb)
    M              = build_incidence(ingredients, recipes)
    C              = build_cooccurrence(ingredients, M)
    row_ord, col_ord = spectral_bicluster_order(M)

    # Filter out tier-0 placeholder recipes (score_final == 0) but keep them
    # tagged so the UI can hide them; don't remove entirely
    for r in recipes:
        r["is_placeholder"] = (r["score_final"] == 0)

    real_recipes = [r for r in recipes if not r["is_placeholder"]]

    data = {
        "meta": {
            "generated":      datetime.now().isoformat(),
            "n_ingredients":  len(ingredients),
            "n_recipes":      len(recipes),
            "n_real_recipes": len(real_recipes),
            "tiers":          sorted(set(r["tier_bonus"] for r in real_recipes)),
            "spectral_method": "svd_fiedler" if HAS_NUMPY else "frequency_fallback",
        },
        "ingredients":    ingredients,
        "recipes":        recipes,
        "incidence":      M,         # 19 × N (all recipes incl. placeholders)
        "cooccurrence":   C,         # 19 × 19
        "row_order":      row_ord,   # ingredient ordering for biclustered matrix
        "col_order":      col_ord,   # recipe ordering for biclustered matrix
    }

    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2))
    print(f"✅ Written: {out_path}")
    print(f"   {len(ingredients)} ingredients, {len(recipes)} recipes ({len(real_recipes)} real)")
    print(f"   Tiers: {data['meta']['tiers']}")
    print()

    # Quick sanity check: verify the 3 畢業 recipes mentioned in CLAUDE.md
    target_names = ["覺醒力量濃湯", "不服輸咖啡沙拉", "土王閃電泡芙",
                    "茂盛焗烤酪梨", "重踏酪梨醬脆片", "採蜜巧克力格子鬆餅"]
    found = {r["name"] for r in recipes}
    for t in target_names:
        status = "✅" if t in found else "❌"
        if t in found:
            r = next(r for r in recipes if r["name"] == t)
            ings = ", ".join(f'{i["name"]}×{i["count"]}' for i in r["ingredients"])
            print(f"{status} {t} [{r['tier_bonus']:.0%}] → {ings}")
        else:
            print(f"{status} {t} — NOT FOUND")


if __name__ == "__main__":
    args = sys.argv[1:]
    if "--out" in args:
        out_path = Path(args[args.index("--out") + 1])
    else:
        out_path = OUT
    main(out_path)
